from PIL import Image
import pytesseract, os, cv2, re

# Converts to greyscale
image = cv2.imread('file.JPG')
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
cv2.imwrite('file.JPG', gray)

pytesseract.pytesseract.tesseract_cmd = r"tesseract.exe"

# runs OC engine 
def ocr_core(filename):
    text = pytesseract.image_to_string(Image.open(filename), lang = 'eng')
    return text

Folder = ''
Path = 'Capture.JPG'

SearchMe = ocr_core(Path)

print (SearchMe)

#x = SearchMe.find("MEP")
#POnum = SearchMe[x:x+10][-5:]
POnum = "".join(re.findall(r"(MEP+[^\s]+)",SearchMe))[-5:]

#y = SearchMe.find("Supplier")
#VID = SearchMe[y:y+13].split(' ')[1]
#if "o" in VID:
#    VID = VID.replace("o","0")

#elif "O" in VID:
#    VID = VID.replace("O","0")

VID = "".join(re.findall(r"Supplier+.*",SearchMe))[-4:]

#z = SearchMe.find("Project Number")
#JID = (SearchMe[z:z+28].split(' ')[2])[-5:]

JID = "".join(re.findall(r"(Project Number+.*)",SearchMe))[-5:]

print (POnum, VID, JID)

import openpyxl

wb = openpyxl.load_workbook('test.xlsx', data_only = True )
sheet = wb['Index']

# Return Supplier name as per code
for row in range(1, sheet.max_row + 1):
    for column in "G":  # Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        if sheet[cell_name].value == VID:
            vename = sheet.cell(column = 8, row = row).value[:15]


NewPath = str(Folder+POnum+' '+vename+' - '+JID+'.jpg')

print (POnum, vename, JID)
print (NewPath)

#os.rename(Path,NewPath)